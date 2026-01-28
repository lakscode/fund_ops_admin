"""
Import/Export utility functions for handling XLSX and JSON file operations
"""
import json
import io
from typing import List, Dict, Any, Tuple, Optional, Callable
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def export_to_xlsx(data: List[Dict[str, Any]], fields: List[str], sheet_name: str = "Data") -> io.BytesIO:
    """Export data to XLSX format"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, field in enumerate(fields, 1):
        ws.cell(row=1, column=col_idx, value=field)
        # Auto-adjust column width
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field) + 2, 15)

    # Write data rows
    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(fields, 1):
            value = item.get(field)
            # Convert datetime to string for Excel
            if isinstance(value, datetime):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_json(data: List[Dict[str, Any]], fields: List[str]) -> io.BytesIO:
    """Export data to JSON format"""
    # Filter data to only include specified fields
    filtered_data = []
    for item in data:
        filtered_item = {}
        for field in fields:
            value = item.get(field)
            # Convert datetime to string for JSON
            if isinstance(value, datetime):
                value = value.isoformat()
            filtered_item[field] = value
        filtered_data.append(filtered_item)

    output = io.BytesIO()
    output.write(json.dumps(filtered_data, indent=2, default=str).encode('utf-8'))
    output.seek(0)
    return output


def import_from_xlsx(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Import data from XLSX file
    Returns: (data_rows, errors)
    """
    errors = []
    data = []

    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # Get headers from first row
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header).strip())
            else:
                headers.append(f"column_{col}")

        # Read data rows
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_data = False
            for col_idx, header in enumerate(headers, 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    has_data = True
                row_data[header] = value

            if has_data:
                data.append(row_data)

    except Exception as e:
        errors.append(f"Error reading XLSX file: {str(e)}")

    return data, errors


def import_from_json(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Import data from JSON file
    Returns: (data_rows, errors)
    """
    errors = []
    data = []

    try:
        content = file_content.decode('utf-8')
        parsed = json.loads(content)

        if isinstance(parsed, list):
            data = parsed
        elif isinstance(parsed, dict):
            # If it's a single object, wrap in list
            data = [parsed]
        else:
            errors.append("JSON must be an array or object")

    except json.JSONDecodeError as e:
        errors.append(f"Invalid JSON format: {str(e)}")
    except Exception as e:
        errors.append(f"Error reading JSON file: {str(e)}")

    return data, errors


def validate_and_convert_row(
    row: Dict[str, Any],
    row_idx: int,
    required_fields: List[str],
    field_types: Dict[str, type],
    field_validators: Optional[Dict[str, Callable]] = None
) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    """
    Validate a single row and convert field types
    Returns: (converted_data, errors)
    """
    errors = []
    converted = {}

    # Check required fields
    for field in required_fields:
        if field not in row or row[field] is None or (isinstance(row[field], str) and not row[field].strip()):
            errors.append(f"Row {row_idx}: Missing required field '{field}'")

    if errors:
        return None, errors

    # Convert and validate fields
    for field, value in row.items():
        if value is None or (isinstance(value, str) and not value.strip()):
            converted[field] = None
            continue

        expected_type = field_types.get(field, str)

        try:
            if expected_type == float:
                converted[field] = float(value) if value else None
            elif expected_type == int:
                converted[field] = int(float(value)) if value else None
            elif expected_type == bool:
                if isinstance(value, bool):
                    converted[field] = value
                elif isinstance(value, str):
                    converted[field] = value.lower() in ('true', 'yes', '1', 'active')
                else:
                    converted[field] = bool(value)
            elif expected_type == datetime:
                if isinstance(value, datetime):
                    converted[field] = value
                elif isinstance(value, str):
                    # Try common date formats
                    for fmt in ['%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%m/%d/%Y', '%d/%m/%Y']:
                        try:
                            converted[field] = datetime.strptime(value, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        errors.append(f"Row {row_idx}: Invalid date format for '{field}'")
                        converted[field] = None
            else:
                converted[field] = str(value).strip() if value else None
        except (ValueError, TypeError) as e:
            errors.append(f"Row {row_idx}: Invalid value for '{field}': {value}")
            converted[field] = None

    # Run custom validators
    if field_validators:
        for field, validator in field_validators.items():
            if field in converted and converted[field] is not None:
                is_valid, error_msg = validator(converted[field])
                if not is_valid:
                    errors.append(f"Row {row_idx}: {error_msg}")

    if errors:
        return None, errors

    return converted, []


def generate_template_xlsx(fields: List[str], sample_data: List[Dict[str, Any]], sheet_name: str = "Template") -> io.BytesIO:
    """Generate a template XLSX file with headers and sample data"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, field in enumerate(fields, 1):
        ws.cell(row=1, column=col_idx, value=field)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field) + 2, 15)

    # Write sample data
    for row_idx, item in enumerate(sample_data, 2):
        for col_idx, field in enumerate(fields, 1):
            value = item.get(field)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=col_idx, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_template_json(fields: List[str], sample_data: List[Dict[str, Any]]) -> io.BytesIO:
    """Generate a template JSON file with sample data"""
    # Filter sample data to only include specified fields
    template_data = []
    for item in sample_data:
        filtered_item = {}
        for field in fields:
            value = item.get(field)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d')
            filtered_item[field] = value
        template_data.append(filtered_item)

    output = io.BytesIO()
    output.write(json.dumps(template_data, indent=2, default=str).encode('utf-8'))
    output.seek(0)
    return output
